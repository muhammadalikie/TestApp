import time

import jdatetime

import os

import openpyxl

import sys

from PyQt5.QtWidgets import (

    QApplication, QWidget, QLabel, QLineEdit, QPushButton, QFrame, QTableWidget, QTableWidgetItem, QHeaderView,

    QVBoxLayout, QHBoxLayout, QMessageBox, QButtonGroup, QRadioButton, QScrollArea, QSizePolicy
)

from PyQt5.QtCore import Qt, QTimer

from PyQt5.QtGui import QPixmap, QColor, QCursor


class MainMenuWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("منوی اصلی")
        self.setLayoutDirection(Qt.RightToLeft)
        self.setWindowState(Qt.WindowMaximized)
        self.setStyleSheet("background-color: white;")

        layout = QVBoxLayout()
        layout.setAlignment(Qt.AlignCenter)

        # تصویر
        image_label = QLabel()
        image_label.setPixmap(QPixmap("./Images/MainMenu.png").scaledToHeight(
            1000, Qt.SmoothTransformation))
        image_label.setAlignment(Qt.AlignCenter)

        # عنوان
        title = QLabel("سامانه آزمون")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("font-size: 92px; font-weight: bold; color: #555;")

        # خط زیر عنوان
        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setFrameShadow(QFrame.Sunken)
        line.setStyleSheet("color: #ccc; background-color: #ccc;")
        line.setFixedHeight(2)
        line.setFixedWidth(500)

        # دکمه‌ها
        btn_start = QPushButton("شروع آزمون")
        btn_start.setFixedSize(840, 120)
        btn_start.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                font-size: 52px;
                border-radius: 18px;
            }
            QPushButton:hover {
                background-color: #43A047;  /* ~10% darker */
            }
        """)

        btn_results = QPushButton("مشاهده نتایج")
        btn_results.setFixedSize(840, 120)
        btn_results.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                font-size: 52px;
                border-radius: 18px;
            }
            QPushButton:hover {
                background-color: #43A047;  /* ~10% darker */
            }
        """)

        btn_exit = QPushButton("خروج از برنامه")
        btn_exit.setFixedSize(840, 120)
        btn_exit.setStyleSheet("""
            QPushButton {
                background-color: #C8E6C9;
                color: black;
                font-size: 52px;
                border-radius: 18px;
            }
            QPushButton:hover {
                background-color: #A5D6A7;  /* ~10% darker */
            }
        """)

        btn_start.setCursor(Qt.PointingHandCursor)
        btn_results.setCursor(Qt.PointingHandCursor)
        btn_exit.setCursor(Qt.PointingHandCursor)

        btn_exit.clicked.connect(QApplication.quit)
        btn_start.clicked.connect(self.start_exam)
        btn_results.clicked.connect(self.show_results)

        # اضافه به چیدمان
        layout.addSpacing(-256)  # بالا رفتن تصویر به مقدار ۵۰ پیکسل
        layout.addWidget(image_label, alignment=Qt.AlignCenter)
        layout.addSpacing(124)
        layout.addWidget(title, alignment=Qt.AlignCenter)
        layout.addSpacing(16)
        layout.addWidget(line, alignment=Qt.AlignCenter)
        layout.addSpacing(16)
        layout.addWidget(btn_start, alignment=Qt.AlignCenter)
        layout.addWidget(btn_results, alignment=Qt.AlignCenter)
        layout.addWidget(btn_exit, alignment=Qt.AlignCenter)

        self.setLayout(layout)

    def start_exam(self):

        self.login_window = LoginWindow()

        self.login_window.show()
        self.close()

    def show_results(self):

        file_name = "results.xlsx"

        # اگر فایل اصلاً وجود نداشت
        if not os.path.exists(file_name):
            QMessageBox.information(
                self, "اطلاعاتی یافت نشد", "هنوز هیچ نتیجه‌ای ثبت نشده است.")
            return

        # باز کردن فایل
        wb = openpyxl.load_workbook(file_name)
        ws = wb.active

        # اگر فقط هدر داریم (یعنی فقط یک ردیف)
        if ws.max_row <= 1:
            QMessageBox.information(
                self, "اطلاعاتی یافت نشد", "هنوز هیچ نتیجه‌ای ذخیره نشده است.")
            return

        self.results_window = QWidget()
        self.results_window.setWindowTitle("نتایج ذخیره‌شده")
        self.results_window.setLayoutDirection(Qt.RightToLeft)
        self.results_window.setWindowState(Qt.WindowMaximized)
        self.results_window.setStyleSheet("background-color: white;")

        layout = QVBoxLayout()
        layout.setContentsMargins(40, 40, 40, 40)
        layout.setSpacing(20)

        table = QTableWidget()
        table.setRowCount(ws.max_row - 1)
        table.setColumnCount(ws.max_column)
        table.verticalHeader().setVisible(False)
        table.setHorizontalHeaderLabels([cell.value for cell in ws[1]])
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        table.setStyleSheet("""
            QTableWidget {
                font-size: 46px;
                border: none;  /* ❌ حذف قاب کلی جدول */
                gridline-color: #4CAF50;  /* ✅ خطوط بین سلول‌ها */
            }
            
            QHeaderView::section {
                background-color: #C8E6C9;
                font-weight: bold;
                font-size: 46px;
                padding: 6px;
                border: none;  /* ❌ حذف حاشیه اطراف تیترها */
            }
        """)

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            for j, val in enumerate(row):
                item = QTableWidgetItem(str(val))
                item.setTextAlignment(Qt.AlignCenter)

                # رنگ وضعیت قبولی/عدم قبولی
                if ws[1][j].value == "وضعیت":
                    if str(val).strip() == "قبولی":
                        item.setBackground(QColor("#388E3C"))
                        item.setForeground(Qt.white)
                    elif str(val).strip() == "عدم قبولی":
                        item.setBackground(QColor("#D32F2F"))
                        item.setForeground(Qt.white)

                table.setItem(i, j, item)

        layout.addWidget(table)

        # دکمه بازگشت
        btn_back = QPushButton("بازگشت به منوی اصلی")
        btn_back.setCursor(Qt.PointingHandCursor)
        btn_back.setFixedSize(840, 120)
        btn_back.setStyleSheet("""
            QPushButton {
                background-color: #A5D6A7;
                font-size: 52px;
                border-radius: 18px;
            }
            QPushButton:hover {
                background-color: #81C784;
            }
        """)
        btn_back.clicked.connect(self.back_to_main)

        layout.addWidget(btn_back, alignment=Qt.AlignCenter)
        self.results_window.setLayout(layout)
        self.results_window.show()
        self.close()

    def back_to_main(self):

        self.new_main = MainMenuWindow()

        self.new_main.show()

        self.results_window.close()


class LoginWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("ورود به آزمون")
        self.setLayoutDirection(Qt.RightToLeft)
        self.setWindowState(Qt.WindowMaximized)
        self.setStyleSheet("background-color: white;")
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()
        layout.setAlignment(Qt.AlignCenter)

        # تصویر
        image_label = QLabel()
        image_label.setPixmap(QPixmap("./Images/LoginPage.png").scaledToHeight(
            1000, Qt.SmoothTransformation))
        image_label.setAlignment(Qt.AlignCenter)

        # عنوان
        title_label = QLabel("آزمون آزمایشی")
        title_label.setStyleSheet(
            "font-size: 92px; font-weight: bold; color: #555;")
        title_label.setAlignment(Qt.AlignCenter)

        # خط زیر عنوان
        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setFrameShadow(QFrame.Sunken)
        line.setStyleSheet("color: #ccc; background-color: #ccc;")
        line.setFixedHeight(2)
        line.setFixedWidth(500)

        # فیلد نام کاربری (مثل دکمه طراحی شده)
        self.name_input = QLineEdit()
        # لیبل بالای فیلد نام کاربری
        name_label = QLabel("نام کاربری خود را وارد کنید:")
        name_label.setAlignment(Qt.AlignCenter)
        name_label.setStyleSheet(
            "font-size: 52px; color: #888; margin-bottom: 8px;")

        self.name_input.setFixedSize(840, 120)
        self.name_input.setAlignment(Qt.AlignCenter)
        self.name_input.setStyleSheet("""
            QLineEdit {
                background-color: #C8E6C9;
                border: none;
                border-radius: 18px;
                font-size: 52px;
                padding: 6px;
            }
            
        """)

        # دکمه شروع آزمون
        start_button = QPushButton("شروع آزمون")
        start_button.setCursor(Qt.PointingHandCursor)
        start_button.setFixedSize(840, 120)
        start_button.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                font-size: 52px;
                border-radius: 18px;
            }
            QPushButton:hover {
                background-color: #43A047;
            }
        """)
        start_button.clicked.connect(self.start_exam)

        # افزودن به چیدمان
        layout.addSpacing(-256)
        layout.addWidget(image_label, alignment=Qt.AlignCenter)
        layout.addSpacing(64)
        layout.addWidget(title_label, alignment=Qt.AlignCenter)
        layout.addSpacing(16)
        layout.addWidget(line, alignment=Qt.AlignCenter)
        layout.addSpacing(64)
        layout.addWidget(name_label, alignment=Qt.AlignCenter)
        layout.addWidget(self.name_input, alignment=Qt.AlignCenter)
        layout.addWidget(start_button, alignment=Qt.AlignCenter)

        self.setLayout(layout)

    def start_exam(self):
        username = self.name_input.text().strip()
        if not username:
            QMessageBox.warning(self, "خطا", "لطفاً نام کاربری را وارد کنید.")
            return

        print(f"ورود با نام کاربری: {username}")
        # اتصال به صفحه بعدی آزمون
        self.exam_window = ExamWindow(username)
        self.exam_window.show()
        self.close()


# ########################################################################

# صفحه سوال

# ########################################################################


class QuestionWidget(QWidget):
    def __init__(self, question_data, parent=None):
        super().__init__(parent)
        self.question_data = question_data
        self.correct_index = question_data["correct"]
        self.explanation = question_data["explanation"]
        self.selected_id = -1
        self.init_ui()

    def init_ui(self):
        self.setLayoutDirection(Qt.RightToLeft)
        layout = QVBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setAlignment(Qt.AlignTop)

        # ✅ تصویر سوال (بالای همه)
        image_label = QLabel()
        image_label.setPixmap(
            QPixmap("./Images/question.png").scaledToHeight(1000, Qt.SmoothTransformation))
        image_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(image_label)

        # ✅ سوال
        self.question_label = QLabel(self.question_data["question"])
        self.question_label.setStyleSheet(
            "font-size: 64px; font-weight: bold; color: #444;")
        self.question_label.setAlignment(Qt.AlignCenter)

        layout.addWidget(self.question_label)

        # ✅ گزینه‌ها
        self.options_group = QButtonGroup(self)
        self.option_buttons = []

        for i, text in enumerate(self.question_data["options"]):
            btn = QRadioButton(text)
            btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Minimum)
            btn.setStyleSheet(self.default_option_style())
            btn.setCursor(QCursor(Qt.PointingHandCursor))
            btn.setStyleSheet(self.default_option_style())
            btn.setMinimumHeight(100)
            btn.toggled.connect(self.update_option_styles)
            self.options_group.addButton(btn, i)
            self.option_buttons.append(btn)
            if i == 0:
                layout.addSpacing(64)

            layout.addWidget(btn)

        # ✅ دکمه‌ها
        self.submit_button = QPushButton("ثبت سوال")
        self.submit_button.setCursor(QCursor(Qt.PointingHandCursor))
        self.submit_button.setStyleSheet(self.primary_button_style())
        self.submit_button.setMinimumHeight(100)
        self.submit_button.clicked.connect(self.check_answer)

        self.next_button = QPushButton("سوال بعدی")
        self.next_button.setCursor(QCursor(Qt.PointingHandCursor))
        self.next_button.setStyleSheet(self.primary_button_style())
        self.next_button.setMinimumHeight(100)
        self.next_button.setEnabled(False)

        button_layout = QHBoxLayout()
        button_layout.setSpacing(29)  # فاصله بین دکمه‌ها

        self.submit_button.setSizePolicy(
            QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.next_button.setSizePolicy(
            QSizePolicy.Expanding, QSizePolicy.Fixed)

        button_layout.addWidget(self.submit_button)
        button_layout.addWidget(self.next_button)

        layout.addSpacing(24)
        layout.addLayout(button_layout)

        # ✅ توضیحات
        self.explanation_label = QLabel("")
        self.explanation_label.setWordWrap(True)
        self.explanation_label.setStyleSheet(
            "font-size: 52px; color: #444; margin-top: 24px;")
        self.explanation_label.setAlignment(Qt.AlignCenter)
        self.explanation_label.hide()
        layout.addWidget(self.explanation_label)

        self.setLayout(layout)

    def default_option_style(self):
        return """
            QRadioButton {
                font-size: 46px;
                padding: 20px;
                background-color: #E0E0E0;
                border-radius: 16px;
                font-weight: normal;
                text-align: right;
            }
            QRadioButton::indicator {
            width: 24px;
            height: 24px;
            background: transparent;
            }
        """

    def selected_option_style(self):
        return """
            QRadioButton {
                font-size: 46px;
                padding: 20px;
                background-color: #A5D6A7;
                border-radius: 16px;
                font-weight: bold;
            }
        """

    def correct_option_style(self):
        return """
            QRadioButton {
                background-color: #388E3C;
                color: white;
                font-size: 46px;
                padding: 20px;
                border-radius: 16px;
                font-weight: bold;
            }
        """

    def wrong_option_style(self):
        return """
            QRadioButton {
                background-color: #D32F2F;
                color: white;
                font-size: 46px;
                padding: 20px;
                border-radius: 16px;
                font-weight: bold;
            }
        """

    def primary_button_style(self):
        return """
            QPushButton {
                background-color: #4CAF50;
                color: white;
                font-size: 42px;
                padding: 20px;
                font-weight: bold;
                border-radius: 12px;
            }
            QPushButton:hover {
                background-color: #43A047;
            }
            QPushButton:disabled {
                background-color: #A5D6A7;
                color: white;
            }
        """

    def update_option_styles(self):
        for i, btn in enumerate(self.option_buttons):
            if btn.isChecked():
                btn.setStyleSheet(self.selected_option_style())
                self.selected_id = i
            else:
                btn.setStyleSheet(self.default_option_style())

    def check_answer(self):
        if self.selected_id == -1:
            QMessageBox.warning(self, "توجه", "لطفاً یک گزینه را انتخاب کنید.")
            return

        for i, btn in enumerate(self.option_buttons):
            if i == self.correct_index:
                btn.setStyleSheet(self.correct_option_style())
            elif i == self.selected_id:
                btn.setStyleSheet(self.wrong_option_style())
            else:
                btn.setStyleSheet(self.default_option_style())
            btn.setDisabled(True)

        self.submit_button.setEnabled(False)
        self.next_button.setEnabled(True)
        self.explanation_label.setText(f"📝 توضیح: {self.explanation}")
        self.explanation_label.show()

    def reset(self):
        self.selected_id = -1
        self.explanation_label.hide()
        self.submit_button.setEnabled(True)
        self.next_button.setEnabled(False)
        self.options_group.setExclusive(False)

        for btn in self.option_buttons:
            btn.setChecked(False)
            btn.setEnabled(True)
            btn.setStyleSheet(self.default_option_style())

        self.options_group.setExclusive(True)


class ExamWindow(QWidget):

    def __init__(self, username):

        super().__init__()

        self.username = username

        self.setWindowTitle("آزمون")

        self.setLayoutDirection(Qt.RightToLeft)

        self.setWindowState(Qt.WindowMaximized)

        # نمونه سؤال آزمایشی

        self.questions = [

            {

                "question": "کدام زبان برنامه‌نویسی سطح بالا محسوب می‌شود؟",

                "options": ["اسمبلی", "پایتون", "ماشین", "باینری"],

                "correct": 1,

                "explanation": "پایتون یک زبان سطح بالا و قابل خواندن برای انسان است."

            },

            {

                "question": "کدام یک از موارد زیر زبان نشانه‌گذاری است؟",

                "options": ["HTML", "پایتون", "C++", "جاوا"],

                "correct": 0,

                "explanation": "HTML یک زبان نشانه‌گذاری برای ساخت صفحات وب است."

            },

            {

                "question": "کدام پروتکل برای ارسال ایمیل استفاده می‌شود؟",

                "options": ["HTTP", "FTP", "SMTP", "SSH"],

                "correct": 2,

                "explanation": "پروتکل SMTP برای ارسال ایمیل استفاده می‌شود."

            },

            {

                "question": "کدام یک از موارد زیر مرورگر نیست؟",

                "options": ["Chrome", "Firefox", "Linux", "Edge"],

                "correct": 2,

                "explanation": "لینوکس یک سیستم‌عامل است، نه مرورگر."

            }

        ]

        self.current_index = 0

        self.layout = QVBoxLayout()

        self.setLayout(self.layout)

        self.load_question(self.current_index)

    def load_question(self, index):

        # اگر قبلاً سوالی وجود داره حذفش کن

        if self.layout.count() > 0:

            old_widget = self.layout.itemAt(0).widget()

            if old_widget:

                old_widget.setParent(None)

        # ساخت سوال جدید

        question_data = self.questions[index]

        self.question_widget = QuestionWidget(question_data)

        # وصل کردن دکمه

        if index == len(self.questions) - 1:

            # اگر آخرین سواله، متن دکمه رو تغییر بده

            self.question_widget.next_button.setText("پایان آزمون")

        self.question_widget.next_button.clicked.connect(self.next_question)

        self.layout.addWidget(self.question_widget)

    def next_question(self):

        self.current_index += 1

        if self.current_index < len(self.questions):

            self.load_question(self.current_index)

        else:

            # QMessageBox.information(self, "پایان", "آزمون به پایان رسید.")

            self.finish_exam()

    def finish_exam(self):

        self.final_start_page = FinalExamStartPage(

            self.username, self.start_final_exam)

        self.final_start_page.show()
        self.close()

    def start_final_exam(self, username):

        self.final_exam_window = FinalExamWindow(username)

        self.final_exam_window.show()


class FinalExamStartPage(QWidget):

    def __init__(self, username, on_start_callback):

        super().__init__()

        self.setWindowTitle("شروع آزمون نهایی")

        self.setLayoutDirection(Qt.RightToLeft)

        self.setWindowState(Qt.WindowMaximized)

        self.username = username

        self.on_start_callback = on_start_callback

        layout = QVBoxLayout()

        layout.setAlignment(Qt.AlignCenter)

        label = QLabel(

            "آموزش به پایان رسید. برای شروع آزمون نهایی دکمه زیر را بزنید.")

        label.setStyleSheet("font-size: 20px;")

        label.setAlignment(Qt.AlignCenter)

        start_button = QPushButton("شروع آزمون نهایی")

        start_button.setFixedSize(250, 50)

        start_button.setStyleSheet("font-size: 18px;")

        start_button.clicked.connect(self.start_final_exam)

        layout.addWidget(label)

        layout.addSpacing(30)

        layout.addWidget(start_button)

        self.setLayout(layout)

    def start_final_exam(self):

        self.on_start_callback(self.username)
        self.close()


class FinalExamWindow(QWidget):

    def __init__(self, username):

        super().__init__()

        self.setWindowTitle("آزمون نهایی")

        self.setLayoutDirection(Qt.RightToLeft)

        self.setWindowState(Qt.WindowMaximized)

        self.username = username

        self.current_index = 0

        self.correct_answers = 0

        self.selected_answers = []

        self.start_time = time.time()  # زمان شروع آزمون

        # لیست سوالات آزمون نهایی

        self.questions = [

            {

                "question": "کدام زبان سطح پایین است؟",

                "options": ["پایتون", "C++", "اسمبلی", "جاوا"],

                "correct": 2,

                "explanation": "اسمبلی زبان سطح پایین نزدیک به سخت‌افزار است."

            },

            {

                "question": "وظیفه سیستم‌عامل چیست؟",

                "options": ["طراحی وب", "مدیریت سخت‌افزار", "توسعه بازی", "ویرایش عکس"],

                "correct": 1,

                "explanation": "سیستم‌عامل وظیفه مدیریت منابع سخت‌افزاری و اجرای برنامه‌ها را دارد."

            },

            {

                "question": "کدام‌یک یک مرورگر وب است؟",

                "options": ["ویندوز", "گوگل کروم", "لینوکس", "اندروید"],

                "correct": 1,

                "explanation": "گوگل کروم یکی از مرورگرهای معروف وب است."

            },

            {

                "question": "واحد اندازه‌گیری سرعت پردازنده چیست؟",

                "options": ["وات", "بایت", "گیگاهرتز", "مگاپیکسل"],

                "correct": 2,

                "explanation": "سرعت پردازنده معمولاً بر حسب گیگاهرتز اندازه‌گیری می‌شود."

            },

            {

                "question": "کدام‌یک از انواع حافظه‌ها فرار (Volatile) است؟",

                "options": ["هارد دیسک", "RAM", "DVD", "SSD"],

                "correct": 1,

                "explanation": "RAM حافظه‌ای فرار است که با خاموش شدن سیستم پاک می‌شود."

            },

            {

                "question": "کدام‌یک زبان نشانه‌گذاری است؟",

                "options": ["HTML", "جاوا", "پایتون", "SQL"],

                "correct": 0,

                "explanation": "HTML یک زبان نشانه‌گذاری برای ساخت صفحات وب است."

            },

            {

                "question": "برای اجرای دستورها در سیستم‌عامل ویندوز از چه ابزاری استفاده می‌شود؟",

                "options": ["CMD", "Excel", "Notepad", "Paint"],

                "correct": 0,

                "explanation": "CMD یا Command Prompt ابزار اجرای دستورات متنی در ویندوز است."

            },

            {

                "question": "کدام نوع شبکه برای ارتباط در محدوده‌ی محلی است؟",

                "options": ["LAN", "WAN", "MAN", "PAN"],

                "correct": 0,

                "explanation": "LAN شبکه‌ای برای ارتباط کامپیوترها در یک مکان محدود است."

            },

            {

                "question": "کدام‌یک از کاربردهای Excel است؟",

                "options": ["طراحی سه‌بعدی", "مدیریت بانک اطلاعاتی", "پردازش متن", "محاسبات و جدول‌کشی"],

                "correct": 3,

                "explanation": "Excel برای محاسبات عددی و جدول‌بندی داده‌ها استفاده می‌شود."

            },

            {

                "question": "کدام‌یک از موارد زیر سخت‌افزار نیست؟",

                "options": ["رم", "مادربرد", "مرورگر", "پردازنده"],

                "correct": 2,

                "explanation": "مرورگر یک نرم‌افزار است، نه سخت‌افزار."

            }

        ]

        self.layout = QVBoxLayout()

        self.setLayout(self.layout)

        self.timer_label = QLabel()

        self.timer_label.setAlignment(Qt.AlignLeft)

        self.timer_label.setStyleSheet("font-size: 16px; font-weight: bold;")

        self.layout.addWidget(self.timer_label)

        self.timer = QTimer()

        self.timer.timeout.connect(self.update_timer)

        self.remaining_time = 10 * 60  # 10 دقیقه به ثانیه

        self.timer.start(1000)  # هر ۱ ثانیه یکبار بروزرسانی

        self.load_question(self.current_index)

    def update_timer(self):

        self.remaining_time -= 1

        minutes = self.remaining_time // 60

        seconds = self.remaining_time % 60

        self.timer_label.setText(

            f"⏱ زمان باقی‌مانده: {minutes:02d}:{seconds:02d}")

        if self.remaining_time <= 0:

            self.timer.stop()

            self.finish_exam()

    def load_question(self, index):

        if self.layout.count() > 1:

            old_widget = self.layout.itemAt(1).widget()

            if old_widget:

                old_widget.setParent(None)

        question_data = self.questions[index]

        self.question_widget = QuestionWidget(question_data)

        self.question_widget.next_button.clicked.connect(self.next_question)

        # اگر سوال آخره، متن دکمه بشه پایان آزمون

        if index == len(self.questions) - 1:

            self.question_widget.next_button.setText("پایان آزمون")

        self.layout.addWidget(self.question_widget)

    def next_question(self):

        selected_id = self.question_widget.options_group.checkedId()

        self.selected_answers.append(selected_id)

        if selected_id == self.questions[self.current_index]["correct"]:

            self.correct_answers += 1

        self.current_index += 1

        if self.current_index < len(self.questions):

            self.load_question(self.current_index)

        else:

            self.timer.stop()

            self.finish_exam()

    def finish_exam(self):

        duration = time.time() - self.start_time

        self.result_window = ResultWindow(

            username=self.username,

            correct_count=self.correct_answers,

            total_questions=len(self.questions),

            duration_seconds=duration,

            answers=self.selected_answers,

            questions=self.questions
        )

        self.result_window.show()
        self.close()


class ResultWindow(QWidget):

    def __init__(self, username, correct_count, total_questions, duration_seconds, answers=None, questions=None):

        super().__init__()

        self.setWindowTitle("نتیجه آزمون")

        self.setLayoutDirection(Qt.RightToLeft)

        self.setWindowState(Qt.WindowMaximized)

        self.username = username

        self.answers = answers

        self.questions = questions

        layout = QVBoxLayout()

        layout.setAlignment(Qt.AlignCenter)

        percent = (correct_count / total_questions) * 100

        status = "قبولی" if percent >= 80 else "عدم قبولی"

        duration_min = int(duration_seconds // 60)

        duration_sec = int(duration_seconds % 60)

        # ذخیره در فایل Excel

        self.save_to_excel(username, correct_count,

                           total_questions, percent, status, duration_seconds)

        # نمایش نتیجه

        label_title = QLabel(f"نتیجه آزمون برای: {username}")

        label_title.setStyleSheet("font-size: 24px; font-weight: bold;")

        label_title.setAlignment(Qt.AlignCenter)

        label_score = QLabel(

            f"تعداد پاسخ صحیح: {correct_count} از {total_questions}")

        label_percent = QLabel(f"درصد نهایی: {percent:.1f}%")

        label_time = QLabel(

            f"زمان صرف شده: {duration_min:02d}:{duration_sec:02d}")

        label_status = QLabel(f"وضعیت نهایی: {status}")

        for lbl in [label_score, label_percent, label_time, label_status]:

            lbl.setStyleSheet("font-size: 18px;")

            lbl.setAlignment(Qt.AlignCenter)

        # دکمه رفتن به مرور پاسخ‌ها

        btn_review = QPushButton("مرور پاسخ‌ها")

        btn_review.setFixedSize(200, 50)

        btn_review.setStyleSheet("font-size: 16px;")

        btn_review.clicked.connect(self.open_review_page)

        layout.addWidget(label_title)

        layout.addSpacing(20)

        layout.addWidget(label_score)

        layout.addWidget(label_percent)

        layout.addWidget(label_time)

        layout.addWidget(label_status)

        layout.addSpacing(30)

        layout.addWidget(btn_review)

        self.setLayout(layout)

    import jdatetime

    def save_to_excel(self, username, correct, total, percent, status, duration_sec):
        file_name = "results.xlsx"

        if os.path.exists(file_name):
            wb = openpyxl.load_workbook(file_name)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append([
                "ردیف", "نام کاربری", "تعداد درست", "کل سوالات",
                "درصد", "وضعیت", "زمان آزمون", "تاریخ و ساعت آزمون"
            ])

        row_number = ws.max_row

        minutes = int(duration_sec // 60)
        seconds = int(duration_sec % 60)
        time_str = f"{minutes:02d}:{seconds:02d}"

        # تاریخ شمسی
        now = jdatetime.datetime.now()
        datetime_str = now.strftime("%Y/%m/%d - %H:%M")

        ws.append([
            row_number,
            username,
            correct,
            total,
            f"{percent:.1f}%",
            status,
            time_str,
            datetime_str
        ])

        wb.save(file_name)

    def open_review_page(self):

        self.review_window = ReviewWindow(self.questions, self.answers)

        self.review_window.show()
        self.close()


class ReviewWindow(QWidget):

    def __init__(self, questions, user_answers):

        super().__init__()

        self.setWindowTitle("مرور پاسخ‌ها")

        self.setLayoutDirection(Qt.RightToLeft)

        self.setWindowState(Qt.WindowMaximized)

        # محتوای اسکرول‌شونده

        content_widget = QWidget()

        layout = QVBoxLayout()

        layout.setAlignment(Qt.AlignTop)

        for i, question in enumerate(questions):

            q_label = QLabel(f"{i+1}. {question['question']}")

            q_label.setStyleSheet("font-weight: bold; font-size: 16px;")

            layout.addWidget(q_label)

            correct_index = question["correct"]

            user_index = user_answers[i]

            for j, option in enumerate(question["options"]):

                option_text = f"- {option}"

                if j == correct_index and j == user_index:

                    style = "color: green; font-weight: bold;"

                    option_text += " ✅ (درست انتخاب شده)"

                elif j == correct_index:

                    style = "color: green;"

                    option_text += " ✅ (پاسخ صحیح)"

                elif j == user_index:

                    style = "color: red;"

                    option_text += " ❌ (پاسخ شما)"

                else:

                    style = "color: gray;"

                opt_label = QLabel(option_text)

                opt_label.setStyleSheet(

                    f"font-size: 14px; padding-right: 10px; {style}")

                layout.addWidget(opt_label)

            layout.addSpacing(10)

        # دکمه‌ها

        btn_layout = QHBoxLayout()

        btn_layout.setAlignment(Qt.AlignCenter)

        btn_restart = QPushButton("بازگشت به صفحه ورود")

        btn_restart.setFixedWidth(200)

        btn_restart.clicked.connect(self.go_to_main_menu)

        btn_exit = QPushButton("خروج از برنامه")

        btn_exit.setFixedWidth(150)

        btn_exit.clicked.connect(QApplication.quit)

        btn_layout.addWidget(btn_restart)

        btn_layout.addSpacing(30)

        btn_layout.addWidget(btn_exit)

        layout.addSpacing(30)

        layout.addLayout(btn_layout)

        content_widget.setLayout(layout)

        scroll_area = QScrollArea()

        scroll_area.setWidgetResizable(True)

        scroll_area.setWidget(content_widget)

        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)

        main_layout = QVBoxLayout()

        main_layout.addWidget(scroll_area)

        self.setLayout(main_layout)

    def go_to_main_menu(self):
        self.menu = MainMenuWindow()
        self.menu.show()
        self.close()


if __name__ == "__main__":

    app = QApplication(sys.argv)

    window = MainMenuWindow()

    window.show()

    sys.exit(app.exec_())
